#!/usr/bin/env python
# -*- coding: utf-8 -*- 

import sys
import pickle
import openpyxl
import time

MAX_COL = 100
MAX_ROW = 310
HEADER = 2

MAX_LHS_WIDTH = 3
COURSE_COL = 2
FIRST_NAME_COL = 4
WDAY_COL = 2

WDAYS = {u"воскресенье" : 0, 
         u"понедельник" : 1,
         u"вторник" : 2, 
         u"среда" : 3, 
         u"четверг" : 4, 
         u"пятница" : 5,
         u"суббота" : 6}

LUNCH = u'Обед'
BREAKFAST = u'Завтрак'
MEAL_DB = 'meal.db'

def get_pers_columns(ws, name_query):
    start = time.time()
    pers_columns = {}
    for i in xrange(FIRST_NAME_COL, MAX_COL):
        cell = ws.cell(row = HEADER, column = i).value
        if cell and name_query.lower() in cell.lower():
            pers_columns[i] = cell
    
    print('get_pers_columns', time.time() - start, 'sec')
    
    return pers_columns

def enumerate_persons(ws):
    start = time.time()
    for i in xrange(FIRST_NAME_COL, MAX_COL):
        cell = ws.cell(row = HEADER, column = i).value
        if cell:
            yield (cell.lower(), i)
    
    print('enumerate_persons', time.time() - start, 'sec')

def split_days(ws):
    start = time.time()
    split_rows = {}
    last_split = None
    for i in xrange(1, MAX_ROW):
        cell = ws.cell(row = i, column = WDAY_COL).value
        for wd in WDAYS:
            if cell and wd in unicode(cell).lower():
                if last_split is not None:
                    split_rows[last_split[0]] = (last_split[1], i)
                last_split = (WDAYS[wd], i+1)
                break
                    
    split_rows[last_split[0]] = (last_split[1], MAX_ROW)

    if len(split_rows) > len(WDAYS):
        print split_rows
        raise ValueError("Repeated weekdays detected")
    
    print('split_days', time.time() - start, 'sec')
    
    return split_rows

def get_order(ws, split_rows, day, col):
    start = time.time()
    order = []

    for index, row in enumerate(ws.iter_rows(min_row=split_rows[day][0], min_col=col, max_row=split_rows[day][1], max_col=col)):
        index = split_rows[day][0] + index
        for c in row:
            cell = c.value
            count_mark = str(cell).strip()
            if cell and count_mark:
                course = ws.cell(row = index, column = COURSE_COL).value
                mass = ws.cell(row = index+1, column = 1).value
                num_ord = ws.cell(row = index+1, column = COURSE_COL+1).value
                desc = u""
                if not mass and not num_ord:
                    desc = ws.cell(row = index+1, column = COURSE_COL).value
                    if desc is None:
                        desc = u""
                    
                order.append((course, desc, count_mark))
            
    print('get_order', time.time() - start, 'sec')

    return order

def reply_order(person, wday, hday):
    start = time.time()
    meal = ""
    msg = u''
    # if hday < 12:
    #     msg = BREAKFAST
    #     meal = BREAKFAST
    # else:
    msg = LUNCH
    meal = LUNCH
    with open(MEAL_DB, 'rb') as f:
        meal_db = pickle.load(f)
        if wday not in meal_db[meal]:
            msg = u"No work - no food. That's the law."
        elif person not in meal_db[meal][wday]:
            msg = u"You did not order."
        else:
            msg += "\n" + u"\n".join(meal_db[meal][wday][person])
    f.close()

    print('reply_order', time.time() - start, 'sec')

    return msg

